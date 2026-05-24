from fastapi import APIRouter, Depends, HTTPException, Response
from app.core.security import get_current_user, CoordOrAdmin
from app.db.database import get_supabase
from datetime import datetime
import io

router = APIRouter(prefix="/reports", tags=["Reports"])


# ── Analytics overview ────────────────────────────────────
@router.get("/overview")
async def overview(
    semester: str = None,
    academic_year: str = None,
    current_user=Depends(CoordOrAdmin),
    supabase=Depends(get_supabase),
):
    p_q = supabase.table("placements").select("ojt_status")
    if semester:      p_q = p_q.eq("semester", semester)
    if academic_year: p_q = p_q.eq("academic_year", academic_year)
    placements = p_q.execute().data or []

    status_count = {}
    for p in placements:
        s = p["ojt_status"]
        status_count[s] = status_count.get(s, 0) + 1

    companies = supabase.table("companies").select("is_accredited").execute().data or []
    students  = supabase.table("students").select("program").execute().data or []
    moas      = supabase.table("moa_requests").select("status").execute().data or []

    by_program, moa_status = {}, {}
    for s in students:
        by_program[s["program"]] = by_program.get(s["program"], 0) + 1
    for m in moas:
        moa_status[m["status"]] = moa_status.get(m["status"], 0) + 1

    return {
        "placements": {"total": len(placements), "by_status": status_count},
        "companies":  {"total": len(companies),  "accredited": sum(1 for c in companies if c.get("is_accredited"))},
        "students":   {"total": len(students),   "by_program": by_program},
        "moa":        {"total": len(moas),        "by_status": moa_status},
    }


# ── Students eligible for certificate (completed hours + grade) ──
@router.get("/eligible-for-certificate")
async def eligible_for_certificate(
    semester: str = None,
    academic_year: str = None,
    current_user=Depends(CoordOrAdmin),
    supabase=Depends(get_supabase),
):
    """Returns placements where student has completed hours AND has a final grade."""
    q = supabase.table("placements").select(
        "id, semester, academic_year, ojt_status, "
        "students(id, student_number, required_hours, program, section, "
        "users!students_user_id_fkey(full_name)), "
        "companies(name)"
    )
    if semester:      q = q.eq("semester", semester)
    if academic_year: q = q.eq("academic_year", academic_year)

    placements = q.execute().data or []

    eligible = []
    for p in placements:
        student = p.get("students") or {}
        required = student.get("required_hours", 480)

        # Check rendered hours
        dtr = supabase.table("dtr_logs").select("hours_rendered") \
                  .eq("placement_id", p["id"]).execute()
        rendered = round(sum(r["hours_rendered"] for r in (dtr.data or [])), 1)
        if rendered < required:
            continue

        # Check final grade exists
        grade = supabase.table("ojt_grades").select("final_grade") \
                    .eq("placement_id", p["id"]).execute()
        final_grade = grade.data[0]["final_grade"] if grade.data else None
        if not final_grade:
            continue

        eligible.append({
            "placement_id":  p["id"],
            "student_name":  (student.get("users") or {}).get("full_name", ""),
            "student_number": student.get("student_number", ""),
            "program":       student.get("program", ""),
            "section":       student.get("section", ""),
            "company":       (p.get("companies") or {}).get("name", ""),
            "semester":      p.get("semester", ""),
            "academic_year": p.get("academic_year", ""),
            "rendered_hours": rendered,
            "required_hours": required,
            "final_grade":   final_grade,
        })

    return eligible


# ── Excel export ──────────────────────────────────────────
@router.get("/excel")
async def export_excel(
    semester: str = None,
    academic_year: str = None,
    current_user=Depends(CoordOrAdmin),
    supabase=Depends(get_supabase),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    q = supabase.table("placements").select(
        "*, students(student_number, program, section, required_hours, "
        "users!students_user_id_fkey(full_name)), companies(name)"
    )
    if semester:      q = q.eq("semester", semester)
    if academic_year: q = q.eq("academic_year", academic_year)
    data = q.execute().data or []

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "OJT Summary"

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ctr       = Alignment(horizontal="center", vertical="center")
    thin      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    headers = [
        "Student Number", "Full Name", "Program", "Section",
        "Company", "Semester", "Academic Year",
        "Required Hours", "Rendered Hours", "Remaining Hours",
        "% Complete", "OJT Status", "Final Grade",
    ]

    ws.row_dimensions[1].height = 20
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = ctr; cell.border = thin

    alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")

    for row_idx, p in enumerate(data, 2):
        student = p.get("students") or {}
        company = p.get("companies") or {}

        dtr      = supabase.table("dtr_logs").select("hours_rendered") \
                       .eq("placement_id", p["id"]).execute()
        rendered = round(sum(r["hours_rendered"] for r in (dtr.data or [])), 2)
        required  = student.get("required_hours", 480)
        remaining = max(0, required - rendered)
        pct       = round((rendered / required * 100), 1) if required > 0 else 0.0

        grade_row   = supabase.table("ojt_grades").select("final_grade") \
                          .eq("placement_id", p["id"]).execute()
        final_grade = grade_row.data[0]["final_grade"] if grade_row.data else ""

        row_vals = [
            student.get("student_number", ""),
            (student.get("users") or {}).get("full_name", ""),
            student.get("program", ""), student.get("section", ""),
            company.get("name", ""),
            p.get("semester", ""), p.get("academic_year", ""),
            required, rendered, remaining, f"{pct}%",
            p.get("ojt_status", "").replace("_", " ").title(),
            final_grade,
        ]

        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin
            cell.alignment = Alignment(
                horizontal="center" if col not in (2, 5) else "left",
                vertical="center"
            )
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    col_widths = [16, 28, 12, 10, 28, 12, 14, 14, 14, 15, 12, 14, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output); output.seek(0)

    fname = f"OJT_Summary_{semester or 'ALL'}_{academic_year or 'ALL'}.xlsx"
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── PDF Certificate ───────────────────────────────────────
@router.get("/certificate/{placement_id}")
async def generate_certificate(
    placement_id: str,
    current_user=Depends(get_current_user),
    supabase=Depends(get_supabase),
):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib import colors
        from reportlab.lib.colors import HexColor
    except ImportError:
        raise HTTPException(500, "reportlab not installed")

    result = supabase.table("placements").select(
        "*, students(student_number, program, required_hours, section, "
        "users!students_user_id_fkey(full_name)), companies(name)"
    ).eq("id", placement_id).execute()

    if not result.data:
        raise HTTPException(404, "Placement not found")

    p       = result.data[0]
    student = p["students"]
    company = p["companies"]

    dtr      = supabase.table("dtr_logs").select("hours_rendered") \
                   .eq("placement_id", placement_id).execute()
    rendered = round(sum(r["hours_rendered"] for r in (dtr.data or [])), 1)
    required = student["required_hours"]

    if rendered < required:
        raise HTTPException(400, f"Hours not complete: {rendered}/{required} rendered")

    grade_row   = supabase.table("ojt_grades").select("final_grade") \
                      .eq("placement_id", placement_id).execute()
    final_grade = grade_row.data[0]["final_grade"] if grade_row.data else None
    if not final_grade:
        raise HTTPException(400, "Final grade not yet computed")

    output = io.BytesIO()
    c      = rl_canvas.Canvas(output, pagesize=landscape(A4))
    w, h   = landscape(A4)
    NAVY   = HexColor("#1F4E79")
    GOLD   = HexColor("#C9A84C")

    c.setFillColor(HexColor("#F5F5F5"))
    c.rect(0, 0, w, h, fill=1, stroke=0)
    c.setStrokeColor(NAVY);  c.setLineWidth(10)
    c.rect(1.2*cm, 1.2*cm, w-2.4*cm, h-2.4*cm, fill=0)
    c.setStrokeColor(GOLD);  c.setLineWidth(3)
    c.rect(1.7*cm, 1.7*cm, w-3.4*cm, h-3.4*cm, fill=0)

    c.setFillColor(NAVY);  c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(w/2, h-3.2*cm, "PANGASINAN STATE UNIVERSITY — LINGAYEN CAMPUS")
    c.setFont("Helvetica", 10); c.setFillColor(HexColor("#555555"))
    c.drawCentredString(w/2, h-4*cm, "College of Engineering and Design | OJT Program")
    c.setStrokeColor(GOLD); c.setLineWidth(1.5)
    c.line(4*cm, h-4.6*cm, w-4*cm, h-4.6*cm)

    c.setFillColor(NAVY); c.setFont("Helvetica-Bold", 32)
    c.drawCentredString(w/2, h-6.2*cm, "CERTIFICATE OF COMPLETION")
    c.setFont("Helvetica", 13); c.setFillColor(HexColor("#555555"))
    c.drawCentredString(w/2, h-7.2*cm, "On-the-Job Training Program")

    c.setFont("Helvetica", 12); c.setFillColor(colors.black)
    c.drawCentredString(w/2, h-8.8*cm, "This is to certify that")

    c.setFont("Helvetica-Bold", 24); c.setFillColor(NAVY)
    name = student["users"]["full_name"].upper()
    c.drawCentredString(w/2, h-10.2*cm, name)
    nw = c.stringWidth(name, "Helvetica-Bold", 24)
    c.setStrokeColor(GOLD); c.setLineWidth(1)
    c.line(w/2-nw/2, h-10.5*cm, w/2+nw/2, h-10.5*cm)

    c.setFont("Helvetica", 11); c.setFillColor(colors.black)
    c.drawCentredString(w/2, h-11.5*cm,
        f"Student No: {student['student_number']}   |   Program: {student['program']}   |   Section: {student.get('section','')}")
    c.drawCentredString(w/2, h-12.4*cm,
        "has successfully completed the required On-the-Job Training at")
    c.setFont("Helvetica-Bold", 16); c.setFillColor(NAVY)
    c.drawCentredString(w/2, h-13.5*cm, company["name"].upper())

    c.setFont("Helvetica", 11); c.setFillColor(colors.black)
    c.drawCentredString(w/2, h-14.4*cm,
        f"Hours Rendered: {rendered}   |   Semester: {p.get('semester','')}   |   AY: {p.get('academic_year','')}   |   Final Grade: {final_grade}")

    c.setFont("Helvetica-Oblique", 10); c.setFillColor(HexColor("#777777"))
    c.drawCentredString(w/2, h-15.3*cm, f"Issued on {datetime.now().strftime('%B %d, %Y')}")

    for x, lbl in [(w*0.25, "OJT Coordinator"), (w*0.75, "Campus Director / Dean")]:
        c.setStrokeColor(NAVY); c.setLineWidth(1)
        c.line(x-4*cm, 3.8*cm, x+4*cm, 3.8*cm)
        c.setFont("Helvetica-Bold", 10); c.setFillColor(NAVY)
        c.drawCentredString(x, 3.3*cm, lbl)

    c.save(); output.seek(0)

    return Response(
        content=output.read(),
        media_type="application/pdf",
        headers={"Content-Disposition":
            f"attachment; filename=OJT_Certificate_{student['student_number']}.pdf"},
    )


# ── Semester archive ──────────────────────────────────────
@router.post("/archive")
async def archive_semester(
    semester: str, academic_year: str,
    current_user=Depends(CoordOrAdmin),
    supabase=Depends(get_supabase),
):
    placements = supabase.table("placements").select("ojt_status") \
        .eq("semester", semester).eq("academic_year", academic_year).execute().data or []

    status_count = {}
    for p in placements:
        status_count[p["ojt_status"]] = status_count.get(p["ojt_status"], 0) + 1

    result = supabase.table("semester_archives").insert({
        "semester": semester, "academic_year": academic_year,
        "archived_by": current_user["id"],
        "total": len(placements), "summary": status_count,
    }).execute()

    return {"message": "Semester archived successfully", "archive": result.data[0]}


@router.get("/archives")
async def list_archives(
    current_user=Depends(CoordOrAdmin),
    supabase=Depends(get_supabase),
):
    result = supabase.table("semester_archives").select("*") \
        .order("created_at", desc=True).execute()
    return result.data or []
